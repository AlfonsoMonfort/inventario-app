from flask import Flask, render_template_string, request, redirect, url_for, send_file
from openpyxl import Workbook
from datetime import datetime
import os
from openpyxl import load_workbook

app = Flask(__name__)

inventario = {
    "fecha": "",
    "almacen": "",
    "vendedor": "",
    "articulos": {}
}
# Diccionarios de equivalencias
codigo_a_referencia = {}
referencia_a_descripcion = {}

def cargar_equivalencias():
    wb = load_workbook("equivalencias.xlsx")
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        descripcion, codigo_barras, referencia = row

        if codigo_barras and referencia:
            codigo_a_referencia[str(codigo_barras)] = str(referencia)
            referencia_a_descripcion[str(referencia)] = descripcion

cargar_equivalencias()

HTML = """
<!DOCTYPE html>
<html>
<head>
    <title>Inventario</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <script src="https://cdnjs.cloudflare.com/ajax/libs/quagga/0.12.1/quagga.min.js"></script>
</head>

<body style="font-family: Arial; padding: 20px;">

<h2>Inicio Inventario</h2>
<form method="POST" action="/inicio">
    Almacén:<br>
    <input type="text" name="almacen" required><br><br>

    Vendedor:<br>
    <input type="text" name="vendedor" required><br><br>

    <button type="submit">Empezar</button>
</form>

{% if inventario["fecha"] %}
<hr>

<h2>Escanear Código</h2>

<!-- MENSAJE VISUAL -->
<div id="mensajeEstado" style="
    display:none;
    padding:12px;
    margin-bottom:10px;
    font-weight:bold;
    border-radius:6px;
    text-align:center;">
</div>

Cantidad:<br>
<input type="number" id="cantidad" value="1" min="1"><br><br>

<div id="scanner" style="width:100%; max-width:400px; border:2px solid black;"></div>

<form method="POST" action="/agregar" id="scanForm">
    <input type="hidden" name="codigo" id="codigoInput">
    <input type="hidden" name="cantidad" id="cantidadInput">
</form>

<hr>

<h3>Artículos Escaneados</h3>
<ul>
{% for referencia, cant in inventario["articulos"].items() %}
    <li>
        <b>{{ referencia_a_descripcion.get(referencia, "Artículo no encontrado") }}</b><br>
        Ref: {{ referencia }} — Cantidad: {{ cant }}
    </li>
    <br>
{% endfor %}
</ul>

<br>
<a href="/excel">
    <button style="padding:15px; font-size:16px;">
        Finalizar y Generar Excel
    </button>
</a>

<!-- SONIDOS -->
<!-- Sonido OK (cling suave) -->
<audio id="okSound" src="https://www.soundjay.com/misc/sounds/bell-ringing-05.mp3"></audio>

<!-- Sonido ERROR (beep corto) -->
<audio id="errorSound" src="https://www.soundjay.com/buttons/sounds/beep-07.mp3"></audio>

<script>

let modoDisparo = false;

document.addEventListener("DOMContentLoaded", function() {

    // ------------------------
    // MENSAJE OK / ERROR
    // ------------------------
    let estado = "{{ estado }}";
    let mensaje = document.getElementById("mensajeEstado");

    if (estado === "ok") {
        mensaje.style.display = "block";
        mensaje.style.backgroundColor = "#d4edda";
        mensaje.style.color = "#155724";
        mensaje.innerHTML = "✅ Artículo añadido";
        document.getElementById("okSound").play();
    }

    if (estado === "error") {
        mensaje.style.display = "block";
        mensaje.style.backgroundColor = "#f8d7da";
        mensaje.style.color = "#721c24";
        mensaje.innerHTML = "❌ Código no encontrado";
        document.getElementById("errorSound").play();
    }

    if (estado === "ok" || estado === "error") {
        setTimeout(function() {
            mensaje.style.display = "none";
            window.history.replaceState({}, document.title, "/");
        }, 1000);
    }

    // ------------------------
    // QUAGGA
    // ------------------------

    if (!document.getElementById("scanner")) return;

    Quagga.init({
        inputStream : {
            name : "Live",
            type : "LiveStream",
            target: document.querySelector('#scanner'),
            constraints: {
                facingMode: "environment"
            }
        },
        decoder : {
            readers : ["ean_reader"]
        }
    }, function(err) {
        if (!err) {
            Quagga.start();
        }
    });

    document.getElementById("scanner").addEventListener("click", function() {
        modoDisparo = true;
    });

    Quagga.onDetected(function(result) {

        if (!modoDisparo) return;

        let code = result.codeResult.code;

        if (!/^\d{13}$/.test(code)) return;

        modoDisparo = false;

        document.getElementById("codigoInput").value = code;
        document.getElementById("cantidadInput").value =
            document.getElementById("cantidad").value;

        document.getElementById("scanForm").submit();
    });

});

</script>

{% endif %}

</body>
</html>
"""









@app.route("/")
def home():
    estado = request.args.get("estado")

    return render_template_string(
        HTML,
        inventario=inventario,
        referencia_a_descripcion=referencia_a_descripcion,
        estado=estado
    )



@app.route("/inicio", methods=["POST"])
def inicio():
    inventario["fecha"] = datetime.now().strftime("%Y-%m-%d")
    inventario["almacen"] = request.form["almacen"]
    inventario["vendedor"] = request.form["vendedor"]
    inventario["articulos"] = {}
    return redirect(url_for("home"))

@app.route("/agregar", methods=["POST"])
def agregar():
    codigo = request.form["codigo"]
    cantidad = int(request.form["cantidad"])

    referencia = codigo_a_referencia.get(codigo)

    if not referencia:
        return redirect(url_for("home", estado="error"))

    if referencia in inventario["articulos"]:
        inventario["articulos"][referencia] += cantidad
    else:
        inventario["articulos"][referencia] = cantidad

    return redirect(url_for("home", estado="ok"))


@app.route("/excel")
def excel():
    filename = "inventario.xlsx"

    wb = Workbook()
    ws = wb.active

    # Cabeceras
    ws["A1"] = "fecha"
    ws["B1"] = "almacen"
    ws["C1"] = "referencia"
    ws["D1"] = "cantidad"
    ws["E1"] = "numero_vendedor"
    ws["F1"] = "descripcion"

    fila = 2

    for referencia, cantidad in inventario["articulos"].items():
        ws[f"A{fila}"] = inventario["fecha"]
        ws[f"B{fila}"] = inventario["almacen"]
        ws[f"C{fila}"] = referencia
        ws[f"D{fila}"] = cantidad
        ws[f"E{fila}"] = inventario["vendedor"]
        ws[f"F{fila}"] = referencia_a_descripcion.get(referencia, "")
        fila += 1

    wb.save(filename)

    # 🔥 RESET DEL INVENTARIO
    inventario["fecha"] = ""
    inventario["almacen"] = ""
    inventario["vendedor"] = ""
    inventario["articulos"] = {}

    return send_file(filename, as_attachment=True)








if __name__ == "__main__":
    app.run()
